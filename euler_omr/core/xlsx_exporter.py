"""openpyxl-based grading result exporter."""
from __future__ import annotations
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from euler_omr.models.scan_result import GradeRecord
from euler_omr.core.analysis import AnalysisReport
import structlog

logger = structlog.get_logger(__name__)


class XlsxExporter:
    HEADER_FILL = PatternFill(start_color="2EB891", end_color="2EB891", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="F0F6F6", size=11)
    ALT_FILL = PatternFill(start_color="E8F4F0", end_color="E8F4F0", fill_type="solid")
    THIN_BORDER = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    @staticmethod
    def export(grades: list[GradeRecord], output_path: str, report: AnalysisReport | None = None):
        wb = Workbook()
        ws = wb.active
        ws.title = "Scores"

        # Build headers: Student ID, Version, Q1..QN, Score, Max Score, Percentage
        max_q = max((len(g.answers) for g in grades), default=0)
        base_headers = ["Student ID", "Version"]
        q_headers = [f"Q{i+1}" for i in range(max_q)]
        tail_headers = ["Score", "Max Score", "Percentage"]
        headers = base_headers + q_headers + tail_headers

        for col, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.fill = XlsxExporter.HEADER_FILL
            c.font = XlsxExporter.HEADER_FONT
            c.alignment = Alignment(horizontal="center")
            c.border = XlsxExporter.THIN_BORDER

        for i, g in enumerate(grades, 2):
            col = 1
            ws.cell(row=i, column=col, value=g.student_id).border = XlsxExporter.THIN_BORDER; col += 1
            ws.cell(row=i, column=col, value=g.version).border = XlsxExporter.THIN_BORDER; col += 1
            for q_idx in range(max_q):
                ans = g.answers[q_idx] if q_idx < len(g.answers) else ""
                cell = ws.cell(row=i, column=col, value=ans if ans else "")
                cell.alignment = Alignment(horizontal="center")
                cell.border = XlsxExporter.THIN_BORDER
                col += 1
            ws.cell(row=i, column=col, value=g.score).border = XlsxExporter.THIN_BORDER; col += 1
            ws.cell(row=i, column=col, value=g.max_score).border = XlsxExporter.THIN_BORDER; col += 1
            ws.cell(row=i, column=col, value=g.percentage).border = XlsxExporter.THIN_BORDER

            if i % 2 == 0:
                for c in range(1, len(headers) + 1):
                    ws.cell(row=i, column=c).fill = XlsxExporter.ALT_FILL

        # Column widths
        ws.column_dimensions["A"].width = 16
        ws.column_dimensions["B"].width = 10
        for q_idx in range(max_q):
            col_letter = _col_letter(3 + q_idx)
            ws.column_dimensions[col_letter].width = 5
        score_col = 3 + max_q
        for offset in range(3):
            cl = _col_letter(score_col + offset)
            ws.column_dimensions[cl].width = 14

        # Analysis Summary sheet
        if report:
            ws2 = wb.create_sheet("Analysis Summary")
            summary_headers = ["Metric", "Value"]
            for col, h in enumerate(summary_headers, 1):
                c = ws2.cell(row=1, column=col, value=h)
                c.fill = XlsxExporter.HEADER_FILL
                c.font = XlsxExporter.HEADER_FONT
            rows = [
                ("Overall Mean", report.overall_mean),
                ("Overall Median", report.overall_median),
                ("Overall Mode", report.overall_mode),
                ("Overall Std Dev", report.overall_stddev),
                ("Fairness Verdict", report.fairness_verdict),
            ]
            for vs in report.version_stats:
                rows.append((f"Version {vs.version} Mean", vs.mean))
                rows.append((f"Version {vs.version} Count", vs.count))
            for i, (k, v) in enumerate(rows, 2):
                ws2.cell(row=i, column=1, value=k)
                ws2.cell(row=i, column=2, value=v)
            ws2.column_dimensions["A"].width = 25
            ws2.column_dimensions["B"].width = 18

        wb.save(output_path)
        logger.info("XLSX exported", path=output_path, count=len(grades))


def _col_letter(col_num: int) -> str:
    """Convert 1-based column number to Excel column letter(s)."""
    result = ""
    while col_num > 0:
        col_num, remainder = divmod(col_num - 1, 26)
        result = chr(65 + remainder) + result
    return result
